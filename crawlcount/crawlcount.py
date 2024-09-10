import sys
import time
import csv
from urllib.parse import urljoin, urlparse
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPlainTextEdit,
    QTextEdit,
    QTextBrowser,
    QPushButton,
    QMessageBox,
    QFileDialog,
    QCheckBox
)
from PyQt5.QtGui import QFont, QIcon
from PyQt5.QtCore import Qt
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from collections import Counter, deque
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


class CrawlCount(QWidget):
    def __init__(self):
        super().__init__()
        self.results_data = []  # To store the results for CSV export
        self.initUI()

    def initUI(self):
        # Main vertical layout
        main_layout = QVBoxLayout()

        # Title Label
        # title = QLabel("CrawlCount")
        # title.setFont(QFont("Arial", 18, QFont.Bold))
        # title.setAlignment(Qt.AlignCenter)
        # main_layout.addWidget(title)

        # Create a horizontal layout for the input fields
        input_layout = QHBoxLayout()

        # URLs Label and Textbox
        urls_layout = QVBoxLayout()
        self.urls_label = QLabel("Enter URLs (one per line):")
        self.urls_label.setFont(QFont("Arial", 12))
        urls_layout.addWidget(self.urls_label)

        self.urls_textbox = QPlainTextEdit()
        self.urls_textbox.setFont(QFont("Arial", 12))
        urls_layout.addWidget(self.urls_textbox)

        input_layout.addLayout(urls_layout)

        # Search Terms Label and Textbox
        search_terms_layout = QVBoxLayout()
        self.search_terms_label = QLabel("Enter search terms (one per line):")
        self.search_terms_label.setFont(QFont("Arial", 12))
        search_terms_layout.addWidget(self.search_terms_label)

        self.search_terms_textbox = QTextEdit()
        self.search_terms_textbox.setFont(QFont("Arial", 12))
        search_terms_layout.addWidget(self.search_terms_textbox)
        input_layout.addLayout(search_terms_layout)
        # Add the horizontal layout to the main layout
        main_layout.addLayout(input_layout)

        # Create a horizontal layout for the buttons
        button_layout = QHBoxLayout()

        # Scrape Button
        self.scrape_button = QPushButton("Start scrape")
        self.scrape_button.setFont(QFont("Arial", 14, QFont.Bold))
        self.scrape_button.clicked.connect(self.scrape_and_count)
        self.scrape_button.setToolTip(
            "Scrapes the URLs and searches for each of the terms in the web page."
        )
        button_layout.addWidget(self.scrape_button)

        # Deep Crawl Button
        self.deep_crawl_button = QPushButton("Start deep crawl")
        self.deep_crawl_button.setFont(QFont("Arial", 14, QFont.Bold))
        self.deep_crawl_button.clicked.connect(self.deep_crawl)
        self.deep_crawl_button.setToolTip(
            "The crawler follows all links on the initial page and fetches those linked pages. "
            + "This means the crawler will search all pages that are directly linked from the initial page. "
            + "This may take some time if there are a lot of pages to crawl."
        )
        button_layout.addWidget(self.deep_crawl_button)

        # Upload CSV Button
        self.upload_button = QPushButton("Upload CSV")
        self.upload_button.setFont(QFont("Arial", 14, QFont.Bold))
        self.upload_button.setToolTip(
            "Upload a CSV file with columns 'urls' and 'terms'."
        )
        self.upload_button.clicked.connect(self.upload_csv)
        button_layout.addWidget(self.upload_button)

        # Download Excel Button
        self.download_button = QPushButton("Download Excel")
        self.download_button.setFont(QFont("Arial", 14, QFont.Bold))
        self.download_button.clicked.connect(self.download_excel)
        self.download_button.setToolTip("Download the results to an Excel file.")
        button_layout.addWidget(self.download_button)

        # Download CSV Button
        self.download_button = QPushButton("Download CSV")
        self.download_button.setFont(QFont("Arial", 14, QFont.Bold))
        self.download_button.clicked.connect(self.download_csv)
        self.download_button.setToolTip("Download the results to a CSV file.")
        button_layout.addWidget(self.download_button)

        # Help Button
        self.help_button = QPushButton("Help")
        self.help_button.setFont(QFont("Arial", 14, QFont.Bold))
        self.help_button.clicked.connect(self.show_help)
        self.help_button.setToolTip("Show help information.")
        button_layout.addWidget(self.help_button)

        # Add the button layout to the main layout
        main_layout.addLayout(button_layout)

        # Status Label
        self.status_label = QLabel("Status: Ready")
        self.status_label.setFont(QFont("Arial", 12))
        main_layout.addWidget(self.status_label)

        # Results Label and Scroll Area
        self.results_label = QLabel("Results:")
        self.results_label.setFont(QFont("Arial", 12))
        main_layout.addWidget(self.results_label)

        self.results_textbox = QTextBrowser()
        self.results_textbox.setFont(QFont("Arial", 12))
        self.results_textbox.setReadOnly(True)
        self.results_textbox.setOpenExternalLinks(True)
        main_layout.addWidget(self.results_textbox)

        # Add checkbox to toggle use Selenium instead of Requests
        self.checkbox = QCheckBox("Use Selenium for scraping - uses Requests by default")
        self.checkbox.setToolTip("Slower but more reliable.")
        self.checkbox.setFont(QFont("Arial", 12))
        main_layout.addWidget(self.checkbox)

        # Set the layout and window properties
        self.setLayout(main_layout)
        self.setWindowTitle("CrawlCount | Scrape links to count search terms")
        self.setWindowIcon(QIcon("logo.png"))
        self.setGeometry(100, 100, 800, 600)

        # Apply styles
        self.apply_styles()

        self.show()

    def apply_styles(self):
        # Apply a stylesheet for modern look and feel
        self.setStyleSheet(
            """
            QWidget {
                background-color: #f5f5f5;
            }
            QLabel {
                color: #333;
            }
            QTextEdit, QPlainTextEdit {
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 8px;
                background-color: #ffffff;
                color: #333;
            }
            QPushButton {
                background-color: #6C757D;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #5A6268;
            }
        """
        )

    def scrape_and_count(self):
        self.results_label.setText("Results:")
        timer_start = time.time()

        urls = self.urls_textbox.toPlainText().strip().splitlines()
        search_terms = self.search_terms_textbox.toPlainText().strip().splitlines()

        if not urls or not search_terms:
            QMessageBox.critical(
                self, "Error", "Please enter at least one URL and one search term."
            )
            return

        self.results_textbox.clear()
        self.results_data.clear()  # Clear previous results

        # Initialize counters
        total_urls = len(urls)
        urls_with_any_terms = 0
        urls_with_all_terms = 0
        processed_urls = set()

        # Update status
        self.status_label.setText("Status: Scraping...")
        QApplication.processEvents()

        # Use selenium to get the HTML content
        chrome_options = Options()
        chrome_options.add_argument("--headless")  # Run headless Chrome
        chrome_options.add_argument("--window-size=450,450")
        service = Service(executable_path=ChromeDriverManager().install())

        driver = webdriver.Chrome(service=service)

        for url in urls:
            if url in processed_urls:
                self.results_textbox.append(f"Skipped duplicate URL: {url}\n")
                continue

            try:
                self.status_label.setText(f"Status: Scraping {url}...")
                QApplication.processEvents()  # Update the UI

                driver.get(url)

                html = driver.page_source

                # Use requests to get the HTML content
                timeout = 4
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                    'Accept-Language': 'en-US,en;q=0.5',
                    'Referer': 'https://www.google.com/',
                    'Connection': 'keep-alive',
                }

                # response = requests.get(url, headers=headers, timeout=timeout)
                # response.raise_for_status()
                
                # Parse the HTML content with BeautifulSoup
                soup = BeautifulSoup(html, "html.parser") # response.text if requests packege used
                text = soup.get_text().lower()

                # Count occurrences of each search term
                term_counts = Counter()
                for term in search_terms:
                    term_counts[term.lower()] += text.count(term.lower())

                # Calculate total matches
                total_matches = sum(term_counts.values())

                # Determine if URL contains any or all search terms
                contains_all_terms = all(term.lower() in text for term in search_terms)
                if any(term_counts.values()):
                    urls_with_any_terms += 1
                if contains_all_terms:
                    urls_with_all_terms += 1

                # Store results for CSV
                self.results_data.append({"URL": url, "Counts": term_counts})

                # Display results with total matches and individual term details
                self.results_textbox.append(
                    f"{total_matches} matches for <a href='{url}'>{url}</a><br>"
                )

                for term, count in term_counts.items():
                    self.results_textbox.append(f"  {term}: {count} occurrences")
                self.results_textbox.append("\n")

                # Add URL to the set of processed URLs
                processed_urls.add(url)
            except requests.exceptions.Timeout as e:
                self.results_textbox.append(f"Failed to retrieve {url}: Timeout. {e}\n\n")
                time.sleep(5)
                continue
            except requests.exceptions.RequestException as e:
                self.results_textbox.append(f"Failed to retrieve {url}: RequestException {e}\n\n")
                time.sleep(5)
                continue
            except requests.exceptions.ConnectionError as e:
                self.results_textbox.append(f"Failed to retrieve {url}: ConnectionError. {e}\n\n")
                time.sleep(5)
                continue
            except requests.exceptions.HTTPError as e:
                self.results_textbox.append(f"Failed to retrieve {url}: HTTPError. {e}\n\n")
                time.sleep(5)
                continue

        # Calculate percentages
        percentage_any_terms = (
            (urls_with_any_terms / total_urls) * 100 if total_urls > 0 else 0
        )
        percentage_all_terms = (
            (urls_with_all_terms / total_urls) * 100 if total_urls > 0 else 0
        )

        # Final status update
        self.status_label.setText("Status: Completed")
        QApplication.processEvents()

        # Stop timer
        elapsed_time = time.time() - timer_start

        # Append summary to results
        self.results_label.setText(
            f"Results: {percentage_any_terms:.1f}% of the URLs contained one or more search terms. " + 
            f"{percentage_all_terms:.1f}% included all the search terms. Elapsed time: {elapsed_time:.2f}s."
        )

    def deep_crawl(self):
        self.results_label.setText("Results:")
        timer_start = time.time()

        urls = self.urls_textbox.toPlainText().strip().splitlines()
        search_terms = self.search_terms_textbox.toPlainText().strip().splitlines()

        if not urls or not search_terms:
            QMessageBox.critical(
                self, "Error", "Please enter at least one URL and one search term."
            )
            return

        self.results_textbox.clear()
        self.results_data.clear()  # Clear previous results

        # Initialize URL queue with (URL, depth) tuples and processed URL set
        url_queue = deque((url, 0) for url in urls)
        processed_urls = set()

        # Update status
        self.status_label.setText("Status: Deep Crawling...")
        QApplication.processEvents()

        while url_queue:
            current_url, depth = url_queue.popleft()
            self.crawl_url(
                current_url,
                search_terms,
                processed_urls,
                url_queue,
                depth
            )
            print("Queue size:", len(url_queue))

        # Final status update
        self.status_label.setText("Status: Completed")
        QApplication.processEvents()

        # Stop timer
        elapsed_time = time.time() - timer_start

        # Append summary to results
        self.results_label.setText(
            "Results: X% of the URLs contained one or more search terms. " + 
            f"X% included all the search terms. Elapsed time: {elapsed_time:.2f}s."
        )

    def crawl_url(self, url, search_terms, processed_urls, url_queue, depth):
        # Skip if the URL has already been processed
        if url in processed_urls:
            return

        # Limit crawl depth to 1 level
        if depth > 1:
            return

        try:
            self.status_label.setText(f"Status: Crawling {url}...")
            QApplication.processEvents()  # Update the UI

            response = requests.get(url)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, "html.parser")
            text = soup.get_text().lower()

            # Count occurrences of each search term
            term_counts = Counter()
            for term in search_terms:
                term_counts[term.lower()] += text.count(term.lower())

            # Calculate total matches
            total_matches = sum(term_counts.values())

            # Store results for CSV
            self.results_data.append({"URL": url, "Counts": term_counts})

            # Display results with total matches and individual term details
            self.results_textbox.append(
                f"{total_matches} matches for <a href='{url}'>{url}</a><br>"
            )

            for term, count in term_counts.items():
                self.results_textbox.append(f"  {term}: {count} occurrences")
            self.results_textbox.append("\n")

            # Add URL to the set of processed URLs
            processed_urls.add(url)

            # Add all links found on this page to the queue with increased depth
            all_links = soup.find_all("a", href=True)
            print(f"Found {len(all_links)} links at depth {depth} on {url}.")

            for link in all_links:
                next_url = urljoin(url, link["href"])
                next_url = (
                    urlparse(next_url)._replace(query="", fragment="").geturl()
                )  # Clean URL

                # Only add new URLs to the queue if they haven't been processed
                if next_url not in processed_urls:
                    url_queue.append((next_url, depth + 1))

        except requests.exceptions.RequestException as e:
            self.results_textbox.append(f"Failed to retrieve {url}: {e}\n\n")

    def download_excel(self):
        if not self.results_data:
            QMessageBox.critical(self, "Error", "No results available to download.")
            return

        # Ask the user for a file name and location to save the Excel file
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Results to Excel",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options,
        )

        if file_path:
            try:
                # Create a new Excel workbook and add sheets
                wb = Workbook()

                # Remove the default sheet if present
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])

                # Summary sheet
                summary_ws = wb.create_sheet(title="Summary")
                summary_ws.append(["Base Domain", "Total Count"])

                # Dictionary to store aggregated counts by base domain
                domain_counts = {}
                for data in self.results_data:
                    url = data["URL"]
                    domain = urlparse(url).netloc
                    total_count = sum(data["Counts"].values())

                    if domain in domain_counts:
                        domain_counts[domain] += total_count
                    else:
                        domain_counts[domain] = total_count

                # Write aggregated counts to the Summary sheet
                for domain, count in domain_counts.items():
                    if not domain.startswith("https://"):
                        domain = "https://" + domain

                    summary_ws.append([domain, count])

                # Create clickable URLs in Summary sheet
                for row in summary_ws.iter_rows(min_row=2, max_col=1):
                    for cell in row:
                        cell.hyperlink = cell.value
                        cell.style = "Hyperlink"

                # Detailed sheet
                detailed_ws = wb.create_sheet(title="Detailed")
                detailed_ws.append(["URL", "Search Term", "Count"])

                # Write detailed counts to the detailed sheet
                for data in self.results_data:
                    url = data["URL"]
                    for term, count in data["Counts"].items():
                        detailed_ws.append([url, term, count])

                # Create clickable URLs in Detailed sheet
                for row in detailed_ws.iter_rows(min_row=2, max_col=1):
                    for cell in row:
                        cell.hyperlink = cell.value
                        cell.style = "Hyperlink"

                # Auto-adjust column widths for readability
                for ws in [summary_ws, detailed_ws]:
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter  # Get the column name
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = max_length + 2
                        ws.column_dimensions[column].width = adjusted_width

                # Save the workbook
                wb.save(file_path)

                QMessageBox.information(
                    self, "Success", f"Results successfully saved to {file_path}"
                )
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save Excel file: {e}")

    def download_csv(self):
        if not self.results_data:
            QMessageBox.critical(self, "Error", "No results available to download.")
            return

        # Ask the user for a file name and location to save the CSV
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Results to CSV",
            "",
            "CSV Files (*.csv);;All Files (*)",
            options=options,
        )

        if file_path:
            try:
                with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
                    fieldnames = ["URL", "Search Term", "Count"]
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                    writer.writeheader()
                    for data in self.results_data:
                        url = data["URL"]
                        for term, count in data["Counts"].items():
                            writer.writerow(
                                {"URL": url, "Search Term": term, "Count": count}
                            )

                QMessageBox.information(
                    self, "Success", f"Results successfully saved to {file_path}"
                )
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save CSV file: {e}")

    def upload_csv(self):
        # Open file dialog to select CSV file
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Open CSV File",
            "",
            "CSV Files (*.csv);;All Files (*)",
            options=options,
        )

        if file_path:
            try:
                urls = []
                search_terms = []

                # Read the CSV file and extract URLs and search terms
                with open(file_path, "r", encoding="utf-8") as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        url = row.get("urls", "").strip()
                        term = row.get("terms", "").strip()
                        if url:
                            urls.append(url)
                        if term:
                            search_terms.append(term)

                # Populate the input fields
                self.urls_textbox.setPlainText("\n".join(urls))
                self.search_terms_textbox.setPlainText("\n".join(search_terms))

                QMessageBox.information(
                    self, "Success", "CSV file successfully loaded."
                )
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load CSV file: {e}")

    def show_help(self):
        # Display help information in a pop-up dialog
        help_text = (
            "CrawlCount Help\n\n"
            "1. Enter URLs: Input the URLs you want to scrape in the 'Enter URLs' textbox. Each URL should be on a new line.\n\n"
            "2. Enter Search Terms: Input the search terms you want to look for in the 'Enter search terms' textbox. Each term should be on a new line.\n\n"
            "3. Start scrape: Click the 'Start scrape' button to begin searching the provided URLs for the specified terms. Results will be displayed in the results area.\n\n"
            "4. Start deep crawl: Click the 'Start deep crawl' button to start crawling initial URLs then all links found on the initial URLs. This will search within these linked pages as well to 1 level deep.\n\n"
            "5. Upload CSV: Click the 'Upload CSV' button to upload a CSV file with columns 'urls' and 'terms' to populate the input fields.\n\n"
            "6. Download results: Click the 'Download Excel' or 'Download CSV' buttons to save the results. The Excel version contains aggregated summary counts by domain and detailed terms by URL.\n\n"
            "7. Use Selenium: Tick the checkbox at the bottom to use Selenium which launches a browser to scrape. More reliable and handles problematic sites better but is slower.\n\n"
            "Thanks again for using CrawlCount.\n\n\n"
            "https://www.gnu.org/licenses/gpl-3.0.en.html\n"
            "https://github.com/shedloadofcode/CrawlCount"
        )
        QMessageBox.information(self, "Help", help_text)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = CrawlCount()

    sys.exit(app.exec_())
